""" ===== SDR++ Spreadsheet to Frequency Manager JSON Converter =====
    ** Public Domain - Do whatever you want with it license. **
    Written by: Steven C. Hageman / November 2024

    Written and tested on,
        Python 3.12
        openpyxl 3.1.2
        SDR++ 1.2.1

    Version History,
        1.0 - 18Nov24 - Initial Release
        1.1 - 22Nov24 - Added more error checking to spreadsheet input
        1.2 - 28Nov24 - Fixed ShowOnWaterfall Issue
        1.3 - 21Dec24 - Added more ShowOnWaterfall Error Checking.

"""
import sys
import json
from collections import Counter
from openpyxl import load_workbook


# $ ===== Helper Methods ======================================================
def get_col_row(s: str):
    head = s.rstrip('0123456789')
    tail = s[len(head):]
    return head, tail


def convert_mode(m: str) -> int:
    try:
        if 'NFM' == m.upper().strip():
            return 0
        if 'WFM' == m.upper().strip():
            return 1
        if 'AM' == m.upper().strip():
            return 2
        if 'DSB' == m.upper().strip():
            return 3
        if 'USB' == m.upper().strip():
            return 4
        if 'CW' == m.upper().strip():
            return 5
        if 'LSB' == m.upper().strip():
            return 6
        if 'RAW' == m.upper().strip():
            return 7
    except:
        return -1

    # If here, there was an error
    return -1


def convert_frequency_units(units: str) -> float:
    try:
        if 'HZ' == units.upper().strip():
            return 1.0
        if 'KHZ' == units.upper().strip():
            return 1000.0
        if 'MHZ' == units.upper().strip():
            return 1_000_000.0
        if 'GHZ' == units.upper().strip():
            return 1_000_000_000.0
    except:
        return -1

    # If here , there was an error
    return -1.0


# $ ===== Program Start Point =================================================

# Get command line parameter
n = len(sys.argv)
try:
    path = sys.argv[1].strip()
    print(f'Converting Spreadsheet: {path}')
except:
    msg = f'''\n\n\nError!\n'''\
          f'''You must specify the XLSX Spreadsheet to convert as a command line parameter.\n'''\
          f'''Fix the error and re-run this program.\n'''
    print(msg)
    sys.exit(-1)


if not (".XLSX" in path.upper().strip()):
    msg = f'''\n\n\nError!\n'''\
          f'''You must specify a "XLSX" Spreadsheet to convert on the command line.\n'''\
          f'''Fix the error and re-run this program.\n'''
    print(msg)
    sys.exit(-2)


# Get Worksheet names for entire Workbook
try:
    workbook = load_workbook(filename=path, read_only=True, data_only=True, rich_text=False)
except:
    msg = f'''\n\n\nError!\n'''\
          f'''Could not find or open the specified input Spreadsheet: {path}.\n'''\
          f'''Fix the error and re-run this program.\n'''
    print(msg)
    sys.exit(-3)


# Create Overall Dictionary and Global Variables
master_dictionary = {'bookmarkDisplayMode': 1, 'lists': {}}
last_section_selection = ''


# Iterate over all the Worksheets
for i, active_sheet in enumerate(workbook.sheetnames):

    # Reset local data
    show_on_waterfall = False
    name = []
    frequency = []
    frequency_units = []
    bandwidth = []
    mode = []
    duplicates = []

    # Activate the current Worksheet
    workbook.active = i
    sheet_obj = workbook.active

    # Get sheet dimensions, really just looking for last row of data
    dimensions = sheet_obj.calculate_dimension()
    dimensions_split = dimensions.split(':')

    end_col, end_row = get_col_row(dimensions_split[1])

    # Read first row, col - Look for Waterfall directive
    first_row = sheet_obj.cell(row=1, column=1)
    first_row_value = first_row.value
    first_row_value = str(first_row_value).upper().strip()

    # Show On Waterfall defaults to False unless it says 'True'
    show_on_waterfall = False
    try:
        if 'WATERFALL' in first_row_value:
            first_row_value_split = first_row_value.split("=")
            if 'TRUE' in first_row_value_split[1].upper():
                show_on_waterfall = True
    except:
        msg = f'''\n\n\nError!\n'''\
              f'''The 'ShowOnWaterfall' parameter on sheet: {active_sheet}.\n'''\
              f'''Was not in the correct form. The correct form is: 'ShowOnWaterfall=True' or 'ShowOnWaterfall=False'.\n'''\
              f'''Fix the error and re-run this program.\n'''
        print(msg)
        sys.exit(-4)

    # Read rows looking for 'Name' directive to find where the actual data starts
    data_start_row_counter = 1
    while data_start_row_counter < int(end_row):
        row_contents = sheet_obj.cell(row=data_start_row_counter, column=1).value
        if 'NAME' in row_contents.upper().strip():
            break

        data_start_row_counter += 1

    # Data starts at next row from 'Name' row
    data_start_row_counter += 1

    if data_start_row_counter > (int(end_row)):
        msg = f'''\n\n\nError!\n'''\
              f'''Could not find the header 'Name' on Worksheet: {active_sheet}.\n'''\
              f'''Fix the error and re-run this program.\n'''
        print(msg)
        sys.exit(-5)


    # Read in each Worksheet line, column by column, and put the results in lists
    for i in range(data_start_row_counter, int(end_row) + 1):

        # Column A
        try:
            name.append((sheet_obj.cell(row=i, column=1).value).strip())    # Column A
        except:
            msg = f'''\n\n\nError!\n'''\
                  f''''Name' error in Column: A, Row: {i} on Worksheet: {active_sheet}.\n'''\
                  f'''Fix the error and re-run this program.\n'''
            print(msg)
            sys.exit(-6)

        # Column B
        try:
            frequency.append(sheet_obj.cell(row=i, column=2).value)         # Column B
        except:
            msg = f'''\n\n\nError!\n'''\
                  f''''Frequency' error in Column: B, Row: {i} on Worksheet: {active_sheet}.\n'''\
                  f'''Fix the error and re-run this program.\n'''
            print(msg)
            sys.exit(-7)

        # Column C
        try:
            frequency_units.append(sheet_obj.cell(row=i, column=3).value)   # Column C
        except:
            msg = f'''\n\n\nError!\n'''\
                  f''''Frequency Units' error in Column: C, Row: {i} on Worksheet: {active_sheet}.\n'''\
                  f'''Fix the error and re-run this program.\n'''
            print(msg)
            sys.exit(-8)

        # Column D
        try:
            bandwidth.append(sheet_obj.cell(row=i, column=4).value)         # Column D
        except:
            msg = f'''\n\n\nError!\n'''\
                  f''''Bandwidth' error in Column: D, Row: {i} on Worksheet: {active_sheet}.\n'''\
                  f'''Fix the error and re-run this program.\n'''
            print(msg)
            sys.exit(-9)

        # Column E
        try:
            mode.append(sheet_obj.cell(row=i, column=5).value)              # Column E
        except:
            msg = f'''\n\n\nError!\n'''\
                  f''''Mode' error in Column: E, Row: {i} on Worksheet: {active_sheet}.\n'''\
                  f'''Fix the error and re-run this program.\n'''
            print(msg)
            sys.exit(-10)


    # Can not have duplicate names in the 'name' column (SDR++ design limitation)
    duplicates = [item for item, count in Counter(name).items() if count > 1]
    if len(duplicates) > 0:
        msg = f'''\n\n\nError!\n'''\
                f'''Found duplicate Name(s): {duplicates} on Worksheet: {active_sheet}.\n'''\
                f'''All names must be unique.\n'''\
                f'''Fix the error(s) and re-run this program.\n'''
        print(msg)
        sys.exit(-11)


    # $ ===== Make the dictionaries - innermost to outermost ==================

    # Bookmark dictionary - bandwidth, frequency, mode
    named_bookmarks_dictionary = {}
    for i, _ in enumerate(frequency):
        parameters = {}

        try:
            parameters['bandwidth'] = int(bandwidth[i])
        except:
            msg = f'''\n\n\nError!\n'''\
                    f'''The Bandwidth for: {name[i]}, in Worksheet: {active_sheet} is invalid.\n'''\
                    f'''Fix the error and re-run this program.\n'''
            print(msg)
            sys.exit(-12)

        frequency_conversion = convert_frequency_units(frequency_units[i])

        if frequency_conversion < 0:
            msg = f'''\n\n\nError!\n'''\
                f'''The Frequency Units: {frequency_units[i]} for Name: {name[i]} on Worksheet: {active_sheet} is invalid.\n'''\
                f'''Valid Frequency Units are: 'Hz', 'kHz', 'MHz' and 'GHz'.\n'''\
                f'''Fix the error and re-run this program.\n'''
            print(msg)
            sys.exit(-13)

        try:
            parameters['frequency'] = int(float(frequency[i] * frequency_conversion))
        except:
            msg = f'''\n\n\nError!\n'''\
                    f'''The Frequency: {frequency[i]} for Name: {name[i]}, in Worksheet: {active_sheet} is invalid.\n'''\
                    f'''Fix the error and re-run this program.\n'''
            print(msg)
            sys.exit(-14)

        mode_int = convert_mode(mode[i])
        if mode_int < 0:
            msg = f'''\n\n\nError!\n'''\
                    f'''The Mode: {mode[i]} for Name: {name[i]}, in Worksheet: {active_sheet} is invalid.\n'''\
                    f'''Fix the error and re-run this program.\n'''
            print(msg)
            sys.exit(-15)

        parameters.update({'mode': mode_int})

        # Make Named bookmark dictionaries
        named_bookmark = {name[i]: parameters}
        named_bookmarks_dictionary.update(named_bookmark)


    # Make bookmarks dictionary
    bookmarks_dictionary = {'bookmarks': named_bookmarks_dictionary}

    # Add the Show on waterfall key/value
    wf_item = {'showOnWaterfall': show_on_waterfall}
    bookmarks_dictionary.update(wf_item)

    # Add the worksheet (section) name
    fixed_sheet_name = active_sheet.strip()
    section_dictionary = {fixed_sheet_name: bookmarks_dictionary}

    master_dictionary['lists'].update(section_dictionary)

    last_section_selection = fixed_sheet_name


# Add a dummy 'SelectedList' - Note: SDR++ Will automatically update this
master_dictionary['selectedList'] = last_section_selection

# $ ===== Print the JSON for saving (and optionally debugging) ================

# Optional debugging
# r = json.dumps(master_dictionary)
# loaded_r = json.loads(r)
# pp = json.dumps(loaded_r)
# out = json.dumps(json.loads(pp), indent=4)
# print(out)

try:
    with open("frequency_manager_config.json", "w") as outfile:
        json.dump(master_dictionary, outfile, indent=4)
except:
    msg = f'''\n\n\nError!\n'''\
          f'''Could not Open or Write to the file: "frequency_manager_config.json".\n'''\
          f'''Fix the error and re-run this program.\n'''
    print(msg)
    sys.exit(-16)


# $ ===== Fini ================================================================
print("\nConversion Completed Successfully.\n")
sys.exit(0)
